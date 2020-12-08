from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import PIL
from pathlib import Path
from os import listdir


def grabQuote(quote_number):

    file_path = str(Path.home()) + "\\Emco\\Brian Kelenc - quotes\\"
    wb = openpyxl.load_workbook(file_path + "quote-"+ str(quote_number) + ".xlsx")
    ws = wb['Sheet1']

    # Set quote number
    # ws['F3'].value = self.quoteNumber.text()
    columns = ['SKU', 'Description', 'Quantity', 'Price', 'Line Price']
    
    ## Create empty dictionary
    rows = { }
    # Iterate through the table and grab the items.

    line_num = 0
 
    for row in range(14,ws.max_row):
        line = []
        if(ws.cell(row,2).value is None):
            break
        #print(ws.cell(row,1).value)

        next_col = 2 
        for x in range(0, len(columns)):
            #ws.cell(column=next_col, row=next_row, value = rows[item][columns[x]])
            cellValue = ws.cell(column=next_col, row=row).value
            line.append(cellValue)

            next_col += 1
        rows[line_num] = {}
        rows[line_num]["SKU"] = line[0] 
        rows[line_num]["Description"] = line[2] 
        rows[line_num]["Quantity"] = line[1] 
        rows[line_num]["Price"] = line[3] 
        rows[line_num]["Line Price"] = line[4] 
        line_num += 1
    return rows

  
