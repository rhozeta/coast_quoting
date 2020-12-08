from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QStackedLayout, QWidget, QLabel, QDialog, QApplication, QMainWindow, QVBoxLayout, QTextEdit, QLineEdit, QCompleter, QInputDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QGridLayout, QMessageBox
from PyQt5.QtGui import *
import sys
import pandas as pd
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import PIL
from pathlib import Path
from os import listdir
import loadQuote as lq



######################################################################################
##                            Coast Water Quote Generator                           ##
##                   Nicholas Mazzotta - nicholasmazzotta@gmail.com                 ##
##                                     October 2020                                 ##    
######################################################################################


    

class MyWindow(QMainWindow):
    def __init__(self):
        super(MyWindow, self).__init__()
        
        self.setWindowTitle("Coast Water Quote Generator")
        self.initUI()
       
    def initUI(self):
        # Set window dimensions
        self.setMinimumWidth(1000)
        self.setMinimumHeight(500)
        
        
        # Logo
        self.logo = QLabel(self)
        self.pixmap = QPixmap('coast_logo.png')
        self.pic = self.pixmap.scaled(50,50, QtCore.Qt.KeepAspectRatio)
        self.logo.setPixmap(self.pic)
        self.logo.setMaximumSize(50,50)
        self.logo.setObjectName('logo')

        # Quote Number Entry
        self.quoteNumberLabel = QLabel(self)
        self.quoteNumberLabel.setText("Quote #:")
        self.quoteNumber = QLineEdit(self)
        

        # Submit button
        self.submit = QtWidgets.QPushButton(self)
        self.submit.setText("Submit")
        self.submit.clicked.connect(self.submitted)
        

        # SKU Search
        self.skuSearchLabel = QLabel(self)
        self.skuSearchLabel.setText('Product:')
        self.skuSearch = QLineEdit(self)
        self.skuList = buildProducts()
        completer = QCompleter(self.skuList, self)
        completer.setCaseSensitivity(0)
        completer.setFilterMode(QtCore.Qt.MatchContains)
        self.skuSearch.setCompleter(completer)
        

        # Quantity Entry
        self.quantityLabel = QLabel(self)
        self.quantityLabel.setText('Quantity:')
        self.quantity = QLineEdit(self)
        self.quantity.setValidator(QIntValidator())

        # Quote Total Label
        self.quoteLabel = QtWidgets.QLabel(self)
        self.quoteLabel.setText("$0.00")
        self.quoteTotalRaw = 0
        self.quoteTotalLabel = QLabel(self)
        self.quoteTotalLabel.setText('Total')
        self.quoteLabelWithMarkup = QtWidgets.QLabel(self)
        self.quoteLabelWithMarkup.setText("$0.00")
        self.quoteTotalLabelWithMarkup = QLabel(self)
        self.quoteTotalLabelWithMarkup.setText('Marked Up')
        self.quoteTotalWithMarkup = 0

        # Freight Total Entry
        #self.freightCost = QLineEdit(self)
        
        # Add SKU to quote button
        self.addSku = QtWidgets.QPushButton(self)
        self.addSku.setText("Add Item")
        self.addSku.clicked.connect(self.addSkuToQuote)

        # Menu Bar
        self.menuBar = self.menuBar()
        self.fileMenu = self.menuBar.addMenu("File")

        self.newAction = QtWidgets.QAction("New", self)
        self.newAction.setShortcut("CTRL+N")
    
        self.saveAction = QtWidgets.QAction("Save", self)
        self.saveAction.setShortcut("CTRL+S")
     
        self.loadAction = QtWidgets.QAction("Load", self)
        self.loadAction.setShortcut("CTRL+L")
        self.loadAction.triggered.connect(self.loadQuote)

             
        self.productAction = QtWidgets.QAction("Edit Products", self)
        self.productAction.setShortcut("CTRL+P")
        self.productAction.triggered.connect(self.editProducts)
        

        self.fileMenu.addAction(self.newAction)
        self.fileMenu.addAction(self.saveAction)
        self.fileMenu.addAction(self.loadAction)
        self.fileMenu.addAction(self.productAction)
        
        self.changesSaved = True

        #Create grid
        def createGrid(self):
          
            wid = QtWidgets.QWidget(self)
            self.setCentralWidget(wid)
            layout = QtWidgets.QGridLayout()
            wid.setLayout(layout)
            
            # Top area
            layout.addWidget(self.logo,6,8,2,2)
            layout.addWidget(self.skuSearchLabel,0,0)
            layout.addWidget(self.skuSearch,0,1)

            layout.addWidget(self.quantityLabel,0,2)
            layout.addWidget(self.quantity,0,3)
            self.quantity.setMaximumWidth(75)

            layout.addWidget(self.addSku,0,4)
    
            self.addSku.setMinimumSize(self.addSku.sizeHint())
            layout.addWidget(self.submit,6,7)
    
            self.submit.setMinimumSize(self.submit.sizeHint())
            layout.addWidget(self.quoteNumberLabel,0,6)
            layout.addWidget(self.quoteNumber,0,7)
            self.quoteNumber.setMaximumWidth(100)
            #layout.addWidget(self.freightCost,1,7)

            # Create quote/SKU area grid
            self.quoteTable = QTableWidget(self)
 
            self.quoteTable.setColumnCount(6)
            self.quoteTable.setHorizontalHeaderLabels(['Delete','SKU', 'Description','Quantity', 'Price', 'Line Price'])
            self.quoteTable.verticalHeader().setDefaultSectionSize(50)
            self.quoteTable.horizontalHeader().setDefaultSectionSize(100)
            self.quoteTable.itemChanged.connect(self.updateCellPrice)

            # Stretch columns to fit
            header = self.quoteTable.horizontalHeader()       
            header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
            header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
            header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)     
            header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
            header.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
            header.setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
            
            layout.addWidget(self.quoteTable,1,0,7,5)

            layout.addWidget(self.quoteTotalLabel,2,6)
            layout.addWidget(self.quoteLabel,3,6)

            layout.addWidget(self.quoteTotalLabelWithMarkup,4,6)
            layout.addWidget(self.quoteLabelWithMarkup,5,6)

        # Draw Grid
        createGrid(self)

    # Add Sku to quote box
    def addSkuToQuote(self):  
        if self.skuSearch.text() != "" and self.quantity.text() != "":
            self.quoteChanged()
            self.quoteTable.blockSignals(True)
            rowPosition = self.quoteTable.rowCount()
            self.quoteTable.insertRow(rowPosition)
        
            self.df_full = getAllProducts()
            # Populate each column
            self.sku = self.df_full['SKU'].loc[self.df_full['SKU'] == self.skuSearch.text().split(":")[0]]
            self.quoteTable.setItem(rowPosition, 1, QTableWidgetItem(self.sku.values[0]))


            self.description = self.df_full['Description'].loc[self.df_full['SKU'] == self.skuSearch.text().split(":")[0]]
            self.descriptionItem = QTableWidgetItem(self.description.values[0])
            self.descriptionItem.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.quoteTable.setItem(rowPosition, 2, self.descriptionItem)

            self.quantityItem = QTableWidgetItem(self.quantity.text())
            self.quantityItem.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter) # change the alignment
            self.quoteTable.setItem(rowPosition, 3, self.quantityItem)
            

            self.price = self.df_full['Price'].loc[self.df_full['SKU'] == self.skuSearch.text().split(":")[0]]
            self.price.fillna("0", inplace=True)

            self.delete = QtWidgets.QPushButton(self)
            self.delete.setText("Remove")
            self.delete.clicked.connect(self.removeRow)
            self.delete.setObjectName('deleteButton')
            self.quoteTable.setCellWidget(rowPosition, 0, self.delete)
        else:
            msg = QMessageBox()
            msg.setWindowTitle("Invalid Input")
            msg.setText("Quantity and SKU cannot be blank.")
            msg.setInformativeText("Please provide a value for both.")
            msg.setStandardButtons(QMessageBox.Ok)
            x = msg.exec_()
            return

        # Grab and calculate item prices
        try:
            if self.price.values[0] != "0":
                priceRaw = Decimal(float(self.price.values[0]))
                self.priceFormatted = str(round(priceRaw, 2))
                self.quoteTable.setItem(rowPosition, 4, QTableWidgetItem(self.priceFormatted))

                self.linePrice = float(float(self.quantity.text()) * float(self.price.values[0]))
                self.linePriceRaw = Decimal(self.linePrice)
                self.linePriceFormatted = str(round(self.linePriceRaw, 2))
                self.quoteTable.setItem(rowPosition, 5, QTableWidgetItem(str(self.linePriceFormatted)))
                self.calculateTotals()
                
            else: 
                self.quoteTable.setItem(rowPosition, 4, QTableWidgetItem(str(self.price.values[0])))
                self.quoteTable.setItem(rowPosition, 5, QTableWidgetItem(str(self.price.values[0])))
        
            self.quoteTable.blockSignals(False)
            self.skuSearch.clear()
            self.quantity.clear()
        except:
            msg = QMessageBox()
            msg.setWindowTitle("No Price Value")
            msg.setText("The SKU has no price on record.")
            msg.setInformativeText("The value for price is blank. Please correct the price in the product editor.")
            msg.setStandardButtons(QMessageBox.Ok)
            x = msg.exec_()
            rowCount = self.quoteTable.currentRow()
            self.quoteTable.blockSignals(True)
            self.quoteTable.setItem(rowCount, 4, QTableWidgetItem(str(0)))
            self.quoteTable.setItem(rowCount, 5, QTableWidgetItem(str(0)))
            self.quoteTable.blockSignals(False)
            
            return

    # Calculate quote totals    
    def calculateTotals(self):
        self.quoteTotal = 0
        for row in range(0,self.quoteTable.rowCount()):
            self.quoteTotal += round(Decimal(self.quoteTable.item(row,5).text()),2)
            #print(self.quoteTable.item(row,4).text())
        self.quoteTotal = Decimal(self.quoteTotal)
        self.quoteTotalWithMarkup = round((self.quoteTotal * Decimal(1.35)),2)
        self.quoteLabel.setText("$" + str(round(self.quoteTotal,2)))
        self.quoteLabelWithMarkup.setText("$" + str(round(self.quoteTotalWithMarkup,2)))

    # Quote selection window for loading quotes
    def loadQuote(self):
        
        selectQuote= QDialog()
        selectQuote.setModal = True
        selectQuote.setMinimumSize(350,500)
        selectQuote.setWindowTitle("Load Quote")
        quotes = grabInitialQuotes()

        grid = QGridLayout()
        selectQuote.setLayout(grid)

        

        self.quoteOpenTable = QTableWidget()
        self.quoteOpenTable.setColumnCount(3)
        self.quoteOpenTable.setHorizontalHeaderLabels(['Load','Quote #','City/Region'])
        self.quoteOpenTable.verticalHeader().setDefaultSectionSize(50)
        self.quoteOpenTable.horizontalHeader().setDefaultSectionSize(100)

        grid.addWidget(self.quoteOpenTable,0,0,3,5)

        rowPosition = self.quoteOpenTable.rowCount()
        x = 0
        for quote in quotes:
            self.quoteOpenTable.insertRow(rowPosition)

            self.load = QtWidgets.QPushButton(self)
            self.load.setText("Load")
            self.load.clicked.connect(self.fillQuote)
            self.load.setObjectName('deleteButton')
            self.quoteOpenTable.setCellWidget(rowPosition, 0, self.load)
            self.quoteOpenTable.setItem(rowPosition, 1, QTableWidgetItem(quotes[x]))
            x+=1
        selectQuote.exec()
        if rowPosition > 0:
            self.loadSkuRow = self.quoteOpenTable.currentRow()
            self.loadSkuNumber = self.quoteOpenTable(self.loadSkuRow,0).text()
        else:
            pass
    # Fill quote in window
    def fillQuote(self):
        if self.changesSaved:
            self.quoteTable.setRowCount(0)
            self.quoteTable.blockSignals(True)
            rowPosition = self.quoteTable.rowCount()
            # Populate each column
            self.loadSkuRow = self.quoteOpenTable.currentRow()
            self.loadSkuNumber = self.quoteOpenTable.item(self.loadSkuRow,1).text()
            quote_num = self.loadSkuNumber
            rows = lq.grabQuote(quote_num)
            for row in rows:
                
                self.quoteTable.insertRow(rowPosition)
                self.delete = QtWidgets.QPushButton(self)
                self.delete.setText("Remove")
                self.delete.clicked.connect(self.removeRow)
                self.delete.setObjectName('deleteButton')
                self.quoteTable.setCellWidget(rowPosition, 0, self.delete)

                self.quoteTable.setItem(rowPosition, 1, QTableWidgetItem(str(rows[row]["SKU"])))
                self.quoteTable.setItem(rowPosition, 2, QTableWidgetItem(str(rows[row]["Description"])))
                self.quoteTable.setItem(rowPosition, 3, QTableWidgetItem(str(rows[row]["Quantity"])))
                self.quoteTable.setItem(rowPosition, 4, QTableWidgetItem(str(rows[row]["Price"])))
                self.quoteTable.setItem(rowPosition, 5, QTableWidgetItem(str(rows[row]["Line Price"])))
                self.updateCellPrice(row)
                self.calculateTotals()
                rowPosition += 1
                self.changesSaved = True
            self.quoteTable.blockSignals(False)
        else: 
            popup = QMessageBox(self)
            popup.setIcon(QMessageBox.Warning)
            popup.setText("The currently open quote has unsaved changes.")
            popup.setInformativeText("Do you want to save your changes?")
            popup.setStandardButtons(QMessageBox.Save   |
                                     QMessageBox.Cancel |
                                     QMessageBox.Discard)
            popup.setDefaultButton(QMessageBox.Save)
    
            answer = popup.exec_()
        
            if answer == QMessageBox.Save:
                self.submitted()
    
            elif answer == QMessageBox.Discard:
                self.changesSaved = True
                self.quoteTable.setRowCount(0)
                self.fillQuote()
    
    def editProducts(self):
        editProduct= QDialog()
        editProduct.setModal = True 
        productEditGrid = QGridLayout()
        editProduct.setLayout(productEditGrid)
        editProduct.setMinimumSize(850,500)
        editProduct.setWindowTitle("Edit Products")

        self.editSkuButton = QtWidgets.QPushButton(self)
        self.editSkuButton.setText("Edit Item")
        self.editSkuButton.clicked.connect(self.setSkuToEditFromSearch)
        self.editSkuLabel = QLabel()
        self.editSkuLabel.setText("Up to date")

        self.currentlyEditing = QLabel()
        self.currentlyEditing.setText("Currently Editing:")
        self.currentlyEditingSku = QLabel()
        self.currentlyEditingSku.setText("New")

        self.skuLabel = QLabel()
        self.skuLabel.setText("SKU: ")
        self.currentSkuBox = QLineEdit()
        self.skuStatus = QLabel()
        self.skuStatus.setText("Up to date")
        self.currentSkuBox.textChanged.connect(self.skuChanged)


        self.descriptionLabel = QLabel()
        self.descriptionLabel.setText("Description: ")
        self.descriptionBox = QLineEdit()
        self.descriptionStatus = QLabel()
        self.descriptionStatus.setText("Up to date")
        self.descriptionBox.textChanged.connect(self.descriptionChanged)

        self.priceLabel = QLabel()
        self.priceLabel.setText("Price: ")
        self.priceBox = QLineEdit()
        self.priceStatus = QLabel()
        self.priceStatus.setText("Up to date")
        self.priceBox.textChanged.connect(self.priceChanged)

        self.productSearch = QLineEdit(self)
        self.productList = buildProducts()
        completer = QCompleter(self.productList, self)
        completer.setCaseSensitivity(0)
        completer.setFilterMode(QtCore.Qt.MatchContains)
        self.productSearch.setCompleter(completer)

        self.submitProductChange = QtWidgets.QPushButton()
        self.submitProductChange.setText("Submit Changes")
        self.submitProductChange.clicked.connect(self.saveProductChanges)

        self.addNewProductButton = QtWidgets.QPushButton()
        self.addNewProductButton.setText("Add New Product")
        self.addNewProductButton.clicked.connect(self.addNewProduct)

        self.clearProductButton = QtWidgets.QPushButton()
        self.clearProductButton.setText("Clear")
        self.clearProductButton.clicked.connect(self.clearProducts)

        
        self.productsTable = QTableWidget()
        self.productsTable.setColumnCount(4)
        self.productsTable.setHorizontalHeaderLabels(['Edit','SKU','Description','Price'])
        self.productsTable.verticalHeader().setDefaultSectionSize(50)
        self.productsTable.horizontalHeader().setDefaultSectionSize(100)
        header = self.productsTable.horizontalHeader()       
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)     
        header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        
        self.products = getAllProducts()
        rowPosition = self.productsTable.rowCount()
        # Loop to populate the list of products
        self.fillProducts()

    
        productEditGrid.addWidget(self.productSearch,0,0,1,6)
        productEditGrid.addWidget(self.productsTable,1,0,9,7)
        productEditGrid.addWidget(self.editSkuButton,0,6,1,1)
        productEditGrid.addWidget(self.currentlyEditing,0,8,1,1)
        productEditGrid.addWidget(self.currentlyEditingSku,0,9,1,2)
        productEditGrid.addWidget(self.skuLabel,1,8,1,1)
        productEditGrid.addWidget(self.currentSkuBox,1,9,1,2)
        productEditGrid.addWidget(self.skuStatus,1,11,1,1)
        productEditGrid.addWidget(self.descriptionLabel,2,8,1,1)
        productEditGrid.addWidget(self.descriptionBox,2,9,1,2)
        productEditGrid.addWidget(self.descriptionStatus,2,11,1,1)
        productEditGrid.addWidget(self.priceLabel,3,8,1,1)
        productEditGrid.addWidget(self.priceBox,3,9,1,2)
        productEditGrid.addWidget(self.priceStatus,3,11,1,1)
        productEditGrid.addWidget(self.submitProductChange,5,8,2,4)
        productEditGrid.addWidget(self.addNewProductButton,6,8,2,4)
        productEditGrid.addWidget(self.clearProductButton,7,8,2,4)
        editProduct.exec()

    def skuChanged(self):
        self.skuStatus.setText("Unsaved Changes")
    def descriptionChanged(self):
        self.descriptionStatus.setText("Unsaved Changes")
    def priceChanged(self):
        self.priceStatus.setText("Unsaved Changes")

    def fillProducts(self):
        self.products = getAllProducts()
        rowPosition = self.productsTable.rowCount()
        # Loop to populate the list of products
        for row in range(0, self.products.shape[0]):
            self.productsTable.insertRow(rowPosition)
           
            
            self.edit = QtWidgets.QPushButton(self)
            self.edit.setText("Edit")
            self.productsTable.setCellWidget(rowPosition, 0, self.edit)
            self.edit.clicked.connect(self.setSkuToEditFromTable)

            self.editSku = self.products['SKU'][row]
            self.productsTable.setItem(rowPosition, 1, QTableWidgetItem(str(self.editSku)))

            self.editDescription = self.products["Description"][row]
            self.productsTable.setItem(rowPosition, 2, QTableWidgetItem(str(self.editDescription)))

            self.editPrice = self.products["Price"][row]
            self.productsTable.setItem(rowPosition, 3, QTableWidgetItem(str(self.editPrice)))
    

    def setProductLabelsToSave(self):
        self.skuStatus.setText("Saved")
        self.descriptionStatus.setText("Saved")
        self.priceStatus.setText("Saved")

    def clearProducts(self):
        self.currentSkuBox.setText("")
        self.descriptionBox.setText("")
        self.priceBox.setText("")
        self.currentlyEditingSku.setText("New")

    def setSkuToEditFromSearch(self):
        if self.productSearch.text() == "":
            msg = QMessageBox()
            msg.setWindowTitle("No Product Selected")
            msg.setText("No Valid Product Selected")
            msg.setInformativeText("Please use the search function or select a product from the list. If you need to create a new product, enter the SKU, description and price then click add.")
            msg.setStandardButtons(QMessageBox.Ok)
            x = msg.exec_()

        else:    
            self.currentSkuBox.blockSignals(True)
            self.descriptionBox.blockSignals(True)
            self.priceBox.blockSignals(True)
            self.skuToEdit = self.productSearch.text().split(":")[0]
            self.currentSkuBox.setText(self.skuToEdit)

            description = self.products['Description'].loc[self.products['SKU'] == self.skuToEdit]
            self.descriptionBox.setText(str(description.values[0]))
            price = self.products['Price'].loc[self.products['SKU'] == self.skuToEdit]
            self.priceBox.setText(str(price.values[0]))

        
            self.currentlyEditingSku.setText(self.skuToEdit)

            self.currentSkuBox.blockSignals(False)
            self.descriptionBox.blockSignals(False)
            self.priceBox.blockSignals(False)

    def setSkuToEditFromTable(self):
        self.currentSkuBox.blockSignals(True)
        self.descriptionBox.blockSignals(True)
        self.priceBox.blockSignals(True)

        row = self.productsTable.currentRow()
        self.skuToEdit = self.productsTable.item(row,1).text()
        self.currentSkuBox.setText(self.skuToEdit)

        description = self.products['Description'].loc[self.products['SKU'] == self.skuToEdit]
        self.descriptionBox.setText(str(description.values[0]))
        price = self.products['Price'].loc[self.products['SKU'] == self.skuToEdit]
    
        self.priceBox.setText(str(price.values[0]))
        self.currentlyEditingSku.setText(self.skuToEdit)

        self.currentSkuBox.blockSignals(False)
        self.descriptionBox.blockSignals(False)
        self.priceBox.blockSignals(False)
    def saveProductChanges(self):
        sku = self.currentSkuBox.text()
        if sku not in str(self.products["SKU"]):
        
            self.products.loc[self.products['SKU'] == sku, "Description"] = self.descriptionBox.text()
            self.products.loc[self.products['SKU'] == sku, "Price"] = self.priceBox.text()
            self.products.loc[self.products['SKU'] == sku, "SKU"] = self.currentSkuBox.text()
          
            self.products.to_csv("product_data_test.csv", index=False)

            self.setProductLabelsToSave()
        else:
            msg = QMessageBox()
            msg.setWindowTitle("SKU Number Exists")
            msg.setText("This SKU Number Already Exists")
            msg.setInformativeText("Please provide a unique SKU value for this product.")
            msg.setStandardButtons(QMessageBox.Ok)
            x = msg.exec_()

    def addNewProduct(self):
        newProductSku = self.currentSkuBox.text()
        newProductDescription = self.descriptionBox.text()
        newProductPrice = self.priceBox.text()
        lastRow = len(self.products)
        sku = self.skuToEdit
        if sku not in str(self.products["SKU"]):
            self.products.loc[lastRow, "SKU"] = newProductSku
            self.products.loc[lastRow, "Description"] = newProductDescription
            self.products.loc[lastRow, "Price"] = newProductPrice

            self.products.to_csv("product_data_test.csv", index=False)
            
            self.setProductLabelsToSave()
            print(self.products.tail())
        else:
            msg = QMessageBox()
            msg.setWindowTitle("SKU Number Exists")
            msg.setText("This SKU Number Already Exists")
            msg.setInformativeText("Please provide a unique SKU value for this product.")
            msg.setStandardButtons(QMessageBox.Ok)
            x = msg.exec_()
    # Update cell prices when new cell added and/or when a column value is changed.
    def updateCellPrice(self,row):
        
        self.rowCount = self.quoteTable.currentRow()
        if self.rowCount >= 0:
            self.quoteTable.blockSignals(True)
            price = float(self.quoteTable.item(self.rowCount,4).text())
            quantity = float(self.quoteTable.item(self.rowCount,3).text())
            self.newTotal = price * quantity
            self.newTotal = str(self.newTotal)
            self.quoteTable.setItem(self.rowCount, 5, QTableWidgetItem(self.newTotal))
            self.quoteTable.blockSignals(False)
            self.quoteChanged()
            print("yes")
        else:
            pass
        self.calculateTotals()      

    # Remove SKU from quote
    def removeRow(self):
        self.quoteTable.removeRow(self.quoteTable.currentRow())
        self.calculateTotals()
        self.quoteChanged()
    
    # Submit button pressed
    def submitted(self):
        
        file_path = str(Path.home()) + "\\Emco\\Brian Kelenc - quotes\\"
        all_files = listdir(file_path)
        quotes = [x for x in all_files if "quote" in x]
        quotes = [x[6:-5] for x in quotes]

        if self.quoteNumber.text() in quotes:
            ## quote number already exists
            print("quote exists")
            self.confirm_overwrite()

    
        else:
            self.saveFile()

    # Confirm overwrite existing file
    def confirm_overwrite(self):
        msg = QMessageBox()
        msg.setWindowTitle("Overwrite Existing Quote")
        msg.setText("Quote number " + self.quoteNumber.text() + " already exists.")
        msg.setInformativeText("Would you like to overwrite the existing quote?")
        msg.setIcon(QMessageBox.Warning)
        goBackButton = msg.addButton("Go Back", QMessageBox.YesRole)    
        overwriteButton = msg.addButton("Overwrite Existing", QMessageBox.AcceptRole)  
        msg.setDefaultButton(goBackButton)
        x = msg.exec_()

        if msg.clickedButton() == overwriteButton:
            self.saveFile()

    # Save File
    def saveFile(self):
        wb = openpyxl.load_workbook('quote_template.xlsx')
        ws = wb['Sheet1']

        # Set quote number
        ws['F3'].value = self.quoteNumber.text()
        columns = ['SKU', 'Description', 'Quantity', 'Price', 'Line Price']
        ## Create empty dictionary
        rows = {
            0: {'SKU':'',
                'Quantity': '',
                'Description': '',
                'Price': '',
                'Line Price': ''
                }
        }
        # Iterate through the table and grab the items.
        file_path = str(Path.home()) + "\\Emco\\Brian Kelenc - quotes\\"
    
        for row in range(0, self.quoteTable.rowCount()):
            line = []
            for col in range(1, self.quoteTable.columnCount()):
                item = self.quoteTable.item(row, col).text()            
                line.append(item)  
            rows[row] = {}
            rows[row]["SKU"] = line[0]
            rows[row]["Quantity"] = line[1]
            rows[row]["Description"] = line[2]
            
            rows[row]["Price"] = line[3]
            rows[row]["Line Price"] = line[4]

        next_row = 14
        for item in rows:
            next_col = 2
            for x in range(0, len(columns)):
                ws.cell(column=next_col, row=next_row, value = rows[item][columns[x]])
                
                next_col += 1
            next_row += 1

                
        
        # header = PIL.Image.open('coast_logo.png')
        #ws.add_image(header, 'A1')
        
        
        wb.save(file_path + 'quote-'+ self.quoteNumber.text() + '.xlsx ') 
        self.changesSaved = True
        self.setWindowTitle("Coast Water Quote Generator - Saved")
    
    # Track if unsaved changes exist
    def quoteChanged(self):
        self.changesSaved = False
        self.setWindowTitle("Coast Water Quote Generator - Unsaved Changes")

    # Handle closure of app
    def closeEvent(self,event):

        if self.changesSaved:
            event.accept()
        else:
            popup = QMessageBox(self)
            popup.setIcon(QMessageBox.Warning)
            popup.setText("The document has been modified")
            popup.setInformativeText("Do you want to save your changes?")
            popup.setStandardButtons(QMessageBox.Save   |
                                     QMessageBox.Cancel |
                                     QMessageBox.Discard)
            popup.setDefaultButton(QMessageBox.Save)
    
            answer = popup.exec_()
        
            if answer == QMessageBox.Save:
                self.submitted()
    
            elif answer == QMessageBox.Discard:
                event.accept()
    
            else:
                event.ignore()
       
# Grab products from CSV file
def getAllProducts():
    df_full = pd.read_csv("product_data_test.csv")
    return (df_full)

# Build products search list
def buildProducts():
    df_sku = getAllProducts()["SKU"].astype(str) + ": " + getAllProducts()["Description"].astype(str)
    return (df_sku.to_list())


def window():
    app = QApplication(sys.argv)
    # CSS Styling
    style = """
        QMainWindow{
            background: #3D405B;
        }

        QLabel#logo{
            padding: 0px;
            margin: 0px;
        }

        QLabel{
            color: #F4F1DE;
        }

        QPushButton#deleteButton{
            margin: 10px;
            border-radius: 2px;
            padding: 4px;
        }

        QPushButton{
            color: #F4F1DE;
            background: #E07A5F;
            border-radius: 1px;
            padding: 4px;
        }
        
        QPushButton:hover{
            color: #F4F1DE;
            background: #81B29A;
            border-radius: 1px;
            padding: 4px;
        }

        QTableWidget{
            background: #F4F1DE;
        }

        QMessageBox{
            background: #3D405B;
        }
        QDialog{
            background: #3D405B;
        }
   
    
    """
    app.setStyleSheet(style)
    win = MyWindow()
    win.show()
    sys.exit(app.exec_())

def grabInitialQuotes():
    file_path = str(Path.home()) + "\\Emco\\Brian Kelenc - quotes\\"
    all_files = listdir(file_path)
    quotes = [x for x in all_files if "quote" in x]
    quotes = [x[6:-5] for x in quotes]
    return quotes

# Main, init window and grab products from getAllProduct() 
if __name__ == "__main__":
    buildProducts()
    grabInitialQuotes()
    window()
    